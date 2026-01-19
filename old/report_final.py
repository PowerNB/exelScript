import pandas as pd
import locale
from pathlib import Path
import glob
import time
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
import os
import subprocess
import platform
import sys
from datetime import datetime
import warnings

# Подавление предупреждений
warnings.filterwarnings('ignore')

def open_file_in_default_app(file_path):
    """Открывает файл с помощью стандартного приложения операционной системы."""
    try:
        file_path = Path(file_path).resolve()
        if not file_path.exists():
            print(f"❌ Файл не найден: {file_path}")
            return
        
        if platform.system() == 'Windows':
            os.startfile(str(file_path))
        elif platform.system() == 'Darwin':
            subprocess.Popen(['open', str(file_path)])
        else:
            subprocess.Popen(['xdg-open', str(file_path)])
        
        print(f"📂 Файл открыт: {file_path}")
    except Exception as e:
        print(f"❌ Не удалось открыть файл: {e}")

# === КОНФИГУРАЦИЯ ===
CONFIG = {
    "SOURCE_PATTERN": "Report*.xlsx",
    "DST_FILE": "Отчёт_по_оборотам_бонусов_ПЛ_ОРТК_2024_2025.xlsx",
    "SHEET_NAME_SOURCE": "ВсеЗаправки",
    "COLS_MAPPING": {
        "Время": "date",
        "Бонусов+": "bonus_plus",
        "Бонусов-": "bonus_minus",
        "Объем": "liters",
        "Основание": "reason"
    },
    "NUMBER_FORMATS": {
        "financial": '#,##0.00',
        "rate": '0.00000000'
    }
}

def setup_locale():
    """Попытка установить русскую локаль для форматирования дат."""
    locale_options = ['ru_RU.UTF-8', 'russian', 'ru_RU', 'Russian_Russia.1251']
    for loc in locale_options:
        try:
            locale.setlocale(locale.LC_TIME, loc)
            print(f"✅ Установлена локаль: {loc}")
            return True
        except (locale.Error, Exception):
            continue
    print("⚠️  Не удалось установить русскую локаль. Месяцы будут на английском.")
    return False

def find_source_files(pattern):
    """Поиск исходных файлов по шаблону."""
    source_files = glob.glob(pattern)
    
    if not source_files:
        # Попробуем найти в поддиректориях
        source_files = glob.glob(f"**/{pattern}", recursive=True)
    
    if not source_files:
        print(f"❌ Файлы не найдены по шаблону: {pattern}")
        print("📁 Текущая директория:", os.getcwd())
        return []
    
    print(f"📁 Найдено файлов: {len(source_files)}")
    for i, file in enumerate(source_files[:5], 1):  # Показываем первые 5 файлов
        print(f"   {i}. {file}")
    
    if len(source_files) > 5:
        print(f"   ... и еще {len(source_files) - 5} файлов")
    
    return source_files

def load_excel_file(file_path, config):
    """Загрузка одного Excel файла."""
    try:
        df = pd.read_excel(
            file_path,
            sheet_name=config["SHEET_NAME_SOURCE"],
            usecols=list(config["COLS_MAPPING"].keys()),
            engine="calamine" if Path(file_path).suffix == '.xlsx' else 'openpyxl'
        ).rename(columns=config["COLS_MAPPING"])
        
        print(f"   ✓ Загружен: {Path(file_path).name} ({len(df)} строк)")
        return df
    except Exception as e:
        print(f"   ✗ Ошибка в {Path(file_path).name}: {str(e)[:100]}...")
        return None

def validate_and_clean_data(df):
    """Валидация и очистка данных."""
    print("🔍 Проверка данных...")
    
    # Проверяем наличие обязательных колонок
    required_cols = ["date", "bonus_plus", "bonus_minus", "liters", "reason"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"⚠️  Отсутствуют колонки: {missing_cols}")
        return None
    
    # Преобразование даты
    df["date"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
    
    # Подсчет пропущенных дат
    missing_dates = df["date"].isna().sum()
    if missing_dates > 0:
        print(f"⚠️  Некорректных дат: {missing_dates}")
    
    # Удаление строк без даты
    df = df.dropna(subset=["date"]).copy()
    
    # Преобразование числовых колонок
    numeric_cols = ["bonus_plus", "bonus_minus", "liters"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # Преобразование основания
    df["reason"] = df["reason"].astype(str).str.strip()
    
    # Добавление периода
    df["period"] = df["date"].dt.to_period("M").dt.to_timestamp()
    
    # Статистика по данным
    print(f"   ✓ Валидных строк: {len(df):,}")
    print(f"   ✓ Период данных: {df['date'].min().strftime('%d.%m.%Y')} - {df['date'].max().strftime('%d.%m.%Y')}")
    print(f"   ✓ Всего начислено: {df['bonus_plus'].sum():,.2f}")
    print(f"   ✓ Всего списано: {df['bonus_minus'].abs().sum():,.2f}")
    
    return df

def load_and_process_data(config):
    """Загружает данные из нескольких XLSX файлов, объединяет и очищает их."""
    source_files = find_source_files(config["SOURCE_PATTERN"])
    if not source_files:
        return None
    
    print("📥 Загрузка файлов...")
    all_data = []
    
    for src_file in tqdm(source_files, desc="Обработка", unit="файл", ascii=True):
        df_temp = load_excel_file(src_file, config)
        if df_temp is not None and not df_temp.empty:
            all_data.append(df_temp)
    
    if not all_data:
        print("❌ Не удалось загрузить данные из файлов")
        return None
    
    # Объединение данных
    df = pd.concat(all_data, ignore_index=True, sort=False)
    print(f"📊 Всего строк объединено: {len(df):,}")
    
    # Валидация и очистка
    df = validate_and_clean_data(df)
    
    return df

def calculate_report(df):
    """Выполняет агрегацию данных и расчет отчета."""
    print("📊 Расчет показателей...")
    
    # Создаем копию для безопасности
    df_clean = df.copy()
    
    # Начисления и литры (только положительные начисления)
    positive_bonus = df_clean[df_clean["bonus_plus"] > 0].copy()
    negative_bonus = df_clean[df_clean["bonus_minus"] < 0].copy()
    
    # Группировка по периоду - начисления и литры с бонусами
    report = positive_bonus.groupby("period").agg({
        "bonus_plus": "sum",
        "liters": "sum"
    }).rename(columns={
        "bonus_plus": "Бонусов начислено",
        "liters": "Продано литров с начислением бонусов"
    })
    
    # ✨ НОВАЯ КОЛОНКА: Продано литров всего (включая без бонусов)
    total_liters = df_clean.groupby("period")["liters"].sum()
    report["Продано литров всего"] = total_liters
    
    # Списания (топливо + сопутка)
    if not negative_bonus.empty:
        total_minus = negative_bonus.groupby("period")["bonus_minus"].sum().abs()
        report["Бонусов списано"] = total_minus
    else:
        report["Бонусов списано"] = 0
    
    # Заполнение отсутствующих значений
    report = report.fillna(0)
    
    # Расчет на 1 литр
    report["На 1 литр начислено бонусов"] = report.apply(
        lambda row: row["Бонусов начислено"] / row["Продано литров с начислением бонусов"] 
        if row["Продано литров с начислением бонусов"] != 0 else 0,
        axis=1
    )
    
    # Форматирование периода с русскими названиями месяцев
    try:
        report.index = report.index.strftime("%B %Y")
    except:
        # Если локаль не русская, используем английские названия
        report.index = report.index.strftime("%B %Y")
    
    # Сброс индекса для записи в Excel
    report = report.reset_index().rename(columns={"index": "Период"})
    
    # Добавляем итоговую строку
    total_row = {
        "Период": "ИТОГО",
        "Бонусов начислено": report["Бонусов начислено"].sum(),
        "Продано литров с начислением бонусов": report["Продано литров с начислением бонусов"].sum(),
        "Продано литров всего": report["Продано литров всего"].sum(),  # ✨ НОВОЕ
        "Бонусов списано": report["Бонусов списано"].sum(),
        "На 1 литр начислено бонусов": report["Бонусов начислено"].sum() / report["Продано литров с начислением бонусов"].sum() 
        if report["Продано литров с начислением бонусов"].sum() != 0 else 0
    }
    
    report = pd.concat([report, pd.DataFrame([total_row])], ignore_index=True)
    
    return report

def get_sheet_name_from_data(df_report):
    """
    Генерирует название листа: "Отчет за {первый месяц} - {последний месяц}"
    """
    if len(df_report) <= 1:  # Если только итоговая строка
        return "Отчет"
    
    # Берем первый и последний периоды (исключая итоговую строку)
    first_period = df_report["Период"].iloc[0]
    last_period = df_report["Период"].iloc[-2]  # Предпоследняя строка перед итогом
    
    sheet_name = f"Отчет за {first_period} - {last_period}"
    
    # Ограничение Excel на длину названия листа (31 символ)
    if len(sheet_name) > 31:
        first_month = first_period.split()[0]
        last_month = last_period.split()[0]
        last_year = last_period.split()[1]
        sheet_name = f"{first_month[:3]}-{last_month[:3]} {last_year}"
    
    return sheet_name

def format_excel_file(file_path, sheet_name, config):
    """
    Применяет форматирование Excel:
    - Автоширина колонок
    - Перенос текста в заголовках
    - Финансовый формат для числовых колонок
    - Выделение итоговой строки
    """
    try:
        wb = load_workbook(file_path)
        
        # Если лист существует
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(f"⚠️  Лист '{sheet_name}' не найден, используется активный лист")
            ws = wb.active
            ws.title = sheet_name[:31]  # Ограничение длины названия
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value or ""))
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Форматирование заголовков (первая строка)
        for cell in ws[1]:
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center',
                shrink_to_fit=True
            )
            # Жирный шрифт для заголовков
            cell.font = cell.font.copy(bold=True)
        
        # ✨ ОБНОВЛЕНО: Финансовый формат для числовых колонок (включая D)
        financial_cols = {
            'B': 'Бонусов начислено', 
            'C': 'Продано литров с начислением бонусов',
            'D': 'Продано литров всего',  # ✨ НОВОЕ
            'E': 'Бонусов списано'
        }
        for col_letter, col_name in financial_cols.items():
            if col_letter in ws.column_dimensions:
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.number_format = config["NUMBER_FORMATS"]["financial"]
        
        # ✨ ОБНОВЛЕНО: Формат для колонки F (была E) - На 1 литр начислено бонусов
        if 'F' in ws.column_dimensions:
            for row in range(2, ws.max_row + 1):
                cell = ws[f"F{row}"]
                cell.number_format = config["NUMBER_FORMATS"]["rate"]
        
        # Выделение итоговой строки
        if ws.max_row > 1:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = cell.font.copy(bold=True)
                cell.fill = cell.fill.copy(fill_type="solid", start_color="FFE0E0E0")
        
        # Сохранение
        wb.save(file_path)
        print("✅ Форматирование применено успешно")
        
    except Exception as e:
        print(f"❌ Ошибка форматирования: {e}")

def create_backup(file_path):
    """Создает резервную копию файла."""
    if Path(file_path).exists():
        backup_name = f"{Path(file_path).stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{Path(file_path).suffix}"
        backup_path = Path(file_path).parent / backup_name
        try:
            Path(file_path).rename(backup_path)
            print(f"📦 Создана резервная копия: {backup_name}")
            return True
        except Exception as e:
            print(f"⚠️  Не удалось создать резервную копию: {e}")
    return False

def main():
    """Основная функция выполнения скрипта."""
    start_time = time.time()
    
    print("=" * 70)
    print("🚀 ФОРМИРОВАНИЕ ОТЧЕТА ПО ОБОРОТАМ БОНУСОВ")
    print("=" * 70)
    
    # Установка локали
    setup_locale()
    
    # Проверка исходных файлов
    if not find_source_files(CONFIG["SOURCE_PATTERN"]):
        print("❌ Программа завершена из-за отсутствия исходных файлов")
        input("Нажмите Enter для выхода...")
        sys.exit(1)
    
    # Загрузка и обработка данных
    df_raw = load_and_process_data(CONFIG)
    if df_raw is None:
        print("❌ Не удалось загрузить данные")
        input("Нажмите Enter для выхода...")
        sys.exit(1)
    
    # Расчет отчета
    df_report = calculate_report(df_raw)
    if df_report.empty:
        print("❌ Не удалось рассчитать отчет")
        input("Нажмите Enter для выхода...")
        sys.exit(1)
    
    # Генерация названия листа
    sheet_name = get_sheet_name_from_data(df_report)
    print(f"📋 Название листа: '{sheet_name}'")
    
    # Проверка и создание резервной копии
    dst_path = Path(CONFIG["DST_FILE"])
    if dst_path.exists():
        create_backup(CONFIG["DST_FILE"])
    
    # Сохранение в Excel
    print(f"📝 Сохранение в '{CONFIG['DST_FILE']}'...")
    try:
        with pd.ExcelWriter(
            CONFIG["DST_FILE"],
            engine="openpyxl",
            mode='w'  # Всегда создаем новый файл для чистоты
        ) as writer:
            # Записываем основной отчет
            df_report.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )
            
            # Дополнительно можно добавить лист с сырыми данными
            if len(df_raw) < 10000:  # Если данных не слишком много
                df_raw_sample = df_raw.head(1000).copy()
                df_raw_sample["date"] = df_raw_sample["date"].dt.strftime("%d.%m.%Y %H:%M")
                df_raw_sample.to_excel(
                    writer,
                    sheet_name="Сырые данные",
                    index=False
                )
        
        print("✅ Файл успешно сохранен")
        
        # Применение форматирования
        format_excel_file(CONFIG["DST_FILE"], sheet_name, CONFIG)
        
        # ✨ ОБНОВЛЕНО: Итоговая статистика с новым показателем
        print("\n" + "=" * 70)
        print("📊 ИТОГОВАЯ СТАТИСТИКА")
        print("=" * 70)
        print(f"📄 Файл: {CONFIG['DST_FILE']}")
        print(f"📋 Лист: {sheet_name}")
        print(f"📊 Периодов в отчете: {len(df_report) - 1}")  # Минус итоговая строка
        print(f"📈 Всего начислено бонусов: {df_report['Бонусов начислено'].iloc[-1]:,.2f}")
        print(f"📉 Всего списано бонусов: {df_report['Бонусов списано'].iloc[-1]:,.2f}")
        print(f"⛽ Продано литров (с бонусами): {df_report['Продано литров с начислением бонусов'].iloc[-1]:,.2f}")
        print(f"⛽ Продано литров (всего): {df_report['Продано литров всего'].iloc[-1]:,.2f}")  # ✨ НОВОЕ
        print(f"⏱️  Время выполнения: {time.time() - start_time:.2f} сек")
        print("=" * 70)
        
        # Открываем файл
        open_file_in_default_app(CONFIG["DST_FILE"])
        
    except Exception as e:
        print(f"❌ Ошибка при сохранении файла: {e}")
        print("Подсказка: Закройте файл Excel, если он открыт")
    
    # Завершение программы
    input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Программа прервана пользователем")
    except Exception as e:
        print(f"\n❌ Непредвиденная ошибка: {e}")
        input("Нажмите Enter для выхода...")