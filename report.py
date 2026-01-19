import pandas as pd
import locale
from pathlib import Path
import glob
import time
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
import os
import subprocess
import platform
import sys
from datetime import datetime
import warnings

# Подавление предупреждений
warnings.filterwarnings('ignore')

def open_file_in_default_app(file_path):
    """Открывает файл с помощью стандартного приложения операционной системы."""
    try:
        file_path = Path(file_path).resolve()
        if not file_path.exists():
            print(f"❌ Файл не найден: {file_path}")
            return
        
        if platform.system() == 'Windows':
            os.startfile(str(file_path))
        elif platform.system() == 'Darwin':
            subprocess.Popen(['open', str(file_path)])
        else:
            subprocess.Popen(['xdg-open', str(file_path)])
        
        print(f"📂 Файл открыт: {file_path}")
    except Exception as e:
        print(f"❌ Не удалось открыть файл: {e}")

# === КОНФИГУРАЦИЯ ===
CONFIG = {
    "SOURCE_PATTERN": "Report*.xlsx",
    "DST_FILE": "Отчёт_по_оборотам_бонусов_ПЛ_ОРТК_2024_2025.xlsx",
    "SHEET_NAME_SOURCE": "ВсеЗаправки",
    "COLS_MAPPING": {
        "Время": "date",
        "Бонусов+": "bonus_plus",
        "Бонусов-": "bonus_minus",
        "Объем": "liters",
        "Основание": "reason",
        "Станция": "azs_number",
        "Марка": "fuel_mark"
    },
    "NUMBER_FORMATS": {
        "financial": '#,##0.00',
        "rate": '0.00000000'
    },
    # === НАСТРОЙКИ ФИЛЬТРАЦИИ ===
    "FILTERS": {
        # Включить/выключить фильтрацию (True - включено, False - выключено)
        "ENABLE_FILTERING": True,
        
        # Марки топлива для ИСКЛЮЧЕНИЯ из анализа (по коду)
        # Раскомментируйте нужные строки для исключения марки
        "EXCLUDE_FUEL_MARKS": [
            # 14,  # Аи-92-К5
            # 15,  # Аи-95-К5
            # 17,  # ДТ-Л-К5
            # 18,  # ГАЗ
            # 19,  # Аи-95-К5PR
            # 21,  # Аи100-К5PR
        ],
        
        # Номера АЗС для ИСКЛЮЧЕНИЯ из анализа
        # Добавьте номера АЗС, которые нужно исключить
        "EXCLUDE_AZS": [
            # 1118, # Пример
        ],
    },
    # Справочник марок топлива для отчетности
    "FUEL_MARKS_DICT": {
        14: "Аи-92-К5",
        15: "Аи-95-К5",
        17: "ДТ-Л-К5",
        18: "ГАЗ",
        19: "Аи-95-К5PR",
        21: "Аи100-К5PR"
    }
}

def setup_locale():
    """Попытка установить русскую локаль для форматирования дат."""
    locale_options = ['ru_RU.UTF-8', 'russian', 'ru_RU', 'Russian_Russia.1251']
    for loc in locale_options:
        try:
            locale.setlocale(locale.LC_TIME, loc)
            print(f"✅ Установлена локаль: {loc}")
            return True
        except (locale.Error, Exception):
            continue
    print("⚠️  Не удалось установить русскую локаль. Месяцы будут на английском.")
    return False

def find_source_files(pattern):
    """Поиск исходных файлов по шаблону."""
    source_files = glob.glob(pattern)
    
    if not source_files:
        source_files = glob.glob(f"**/{pattern}", recursive=True)
    
    if not source_files:
        print(f"❌ Файлы не найдены по шаблону: {pattern}")
        print("📁 Текущая директория:", os.getcwd())
        return []
    
    print(f"📁 Найдено файлов: {len(source_files)}")
    for i, file in enumerate(source_files[:5], 1):
        print(f"   {i}. {file}")
    
    if len(source_files) > 5:
        print(f"   ... и еще {len(source_files) - 5} файлов")
    
    return source_files

def load_excel_file(file_path, config):
    """Загрузка одного Excel файла."""
    try:
        # Получаем список колонок для загрузки
        cols_to_load = list(config["COLS_MAPPING"].keys())
        
        df = pd.read_excel(
            file_path,
            sheet_name=config["SHEET_NAME_SOURCE"],
            usecols=cols_to_load,
            engine="calamine" if Path(file_path).suffix == '.xlsx' else 'openpyxl'
        ).rename(columns=config["COLS_MAPPING"])
        
        tqdm.write(f"   ✓ Загружен: {Path(file_path).name} ({len(df)} строк)")
        return df
    except Exception as e:
        tqdm.write(f"   ✗ Ошибка в {Path(file_path).name}: {str(e)[:100]}...")
        return None

def apply_filters(df, config):
    """
    Применяет фильтры для исключения марок топлива и АЗС.
    Возвращает отфильтрованный DataFrame и статистику фильтрации.
    """
    if not config["FILTERS"]["ENABLE_FILTERING"]:
        print("ℹ️  Фильтрация отключена (ENABLE_FILTERING = False)")
        return df, {"filtered_rows": 0, "filtered_fuel": 0, "filtered_azs": 0}
    
    print("\n" + "="*70)
    print("🔍 ПРИМЕНЕНИЕ ФИЛЬТРОВ")
    print("="*70)
    
    initial_count = len(df)
    stats = {
        "filtered_rows": 0,
        "filtered_fuel": 0,
        "filtered_azs": 0,
        "initial_count": initial_count
    }
    
    # Создаем копию для фильтрации
    df_filtered = df.copy()
    
    # === ФИЛЬТРАЦИЯ ПО МАРКАМ ТОПЛИВА ===
    exclude_marks = config["FILTERS"]["EXCLUDE_FUEL_MARKS"]
    if exclude_marks and "fuel_mark" in df_filtered.columns:
        print("\n🚫 Фильтрация по маркам топлива:")
        
        # Преобразуем в числовой формат
        df_filtered["fuel_mark"] = pd.to_numeric(df_filtered["fuel_mark"], errors="coerce")
        
        # Подсчет строк для исключения
        fuel_mask = df_filtered["fuel_mark"].isin(exclude_marks)
        stats["filtered_fuel"] = fuel_mask.sum()
        
        # Вывод информации о исключаемых марках
        if stats["filtered_fuel"] > 0:
            print(f"   Исключаемые марки:")
            for mark_code in exclude_marks:
                mark_name = config["FUEL_MARKS_DICT"].get(mark_code, f"Неизвестная ({mark_code})")
                count = (df_filtered["fuel_mark"] == mark_code).sum()
                if count > 0:
                    bonus_sum = df_filtered[df_filtered["fuel_mark"] == mark_code]["bonus_plus"].sum()
                    liters_sum = df_filtered[df_filtered["fuel_mark"] == mark_code]["liters"].sum()
                    print(f"      • Марка {mark_code} ({mark_name}):")
                    print(f"        - Строк: {count:,}")
                    print(f"        - Бонусов начислено: {bonus_sum:,.2f}")
                    print(f"        - Литров: {liters_sum:,.2f}")
        
        # Применяем фильтр
        df_filtered = df_filtered[~fuel_mask].copy()
        print(f"   ✅ Отфильтровано по маркам: {stats['filtered_fuel']:,} строк")
    else:
        if not exclude_marks:
            print("\n✓ Фильтрация по маркам топлива не настроена (список пуст)")
    
    # === ФИЛЬТРАЦИЯ ПО НОМЕРАМ АЗС ===
    exclude_azs = config["FILTERS"]["EXCLUDE_AZS"]
    if exclude_azs and "azs_number" in df_filtered.columns:
        print("\n🚫 Фильтрация по номерам АЗС:")
        
        # Преобразуем в числовой формат
        df_filtered["azs_number"] = pd.to_numeric(df_filtered["azs_number"], errors="coerce")
        
        # Подсчет строк для исключения
        azs_mask = df_filtered["azs_number"].isin(exclude_azs)
        stats["filtered_azs"] = azs_mask.sum()
        
        # Вывод информации о исключаемых АЗС
        if stats["filtered_azs"] > 0:
            print(f"   Исключаемые АЗС:")
            for azs_num in exclude_azs:
                count = (df_filtered["azs_number"] == azs_num).sum()
                if count > 0:
                    bonus_sum = df_filtered[df_filtered["azs_number"] == azs_num]["bonus_plus"].sum()
                    liters_sum = df_filtered[df_filtered["azs_number"] == azs_num]["liters"].sum()
                    print(f"      • АЗС №{azs_num}:")
                    print(f"        - Строк: {count:,}")
                    print(f"        - Бонусов начислено: {bonus_sum:,.2f}")
                    print(f"        - Литров: {liters_sum:,.2f}")
        
        # Применяем фильтр
        df_filtered = df_filtered[~azs_mask].copy()
        print(f"   ✅ Отфильтровано по АЗС: {stats['filtered_azs']:,} строк")
    else:
        if not exclude_azs:
            print("\n✓ Фильтрация по АЗС не настроена (список пуст)")
    
    # === ОБЩАЯ СТАТИСТИКА ===
    stats["filtered_rows"] = initial_count - len(df_filtered)
    
    print("\n" + "-"*70)
    print("📊 ИТОГИ ФИЛЬТРАЦИИ:")
    print(f"   • Исходных строк: {initial_count:,}")
    print(f"   • Отфильтровано всего: {stats['filtered_rows']:,} ({stats['filtered_rows']/initial_count*100:.2f}%)")
    print(f"   • Осталось для анализа: {len(df_filtered):,}")
    print("="*70)
    
    return df_filtered, stats

def validate_and_clean_data(df):
    """Валидация и очистка данных."""
    print("\n🔍 Проверка и очистка данных...")
    
    # Проверяем наличие обязательных колонок
    required_cols = ["date", "bonus_plus", "bonus_minus", "liters", "reason"]
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        print(f"⚠️  Отсутствуют колонки: {missing_cols}")
        return None
    
    # Преобразование даты
    df["date"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
    
    # Подсчет пропущенных дат
    missing_dates = df["date"].isna().sum()
    if missing_dates > 0:
        print(f"⚠️  Некорректных дат: {missing_dates}")
    
    # Удаление строк без даты
    df = df.dropna(subset=["date"]).copy()
    
    # Преобразование числовых колонок
    numeric_cols = ["bonus_plus", "bonus_minus", "liters"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # Преобразование основания
    df["reason"] = df["reason"].astype(str).str.strip()
    
    # Добавление периода
    df["period"] = df["date"].dt.to_period("M").dt.to_timestamp()
    
    # Статистика по данным
    print(f"   ✓ Валидных строк: {len(df):,}")
    print(f"   ✓ Период данных: {df['date'].min().strftime('%d.%m.%Y')} - {df['date'].max().strftime('%d.%m.%Y')}")
    print(f"   ✓ Всего начислено: {df['bonus_plus'].sum():,.2f}")
    print(f"   ✓ Всего списано: {df['bonus_minus'].abs().sum():,.2f}")
    
    return df

def load_and_process_data(config):
    """Загружает данные из нескольких XLSX файлов, объединяет и очищает их."""
    source_files = find_source_files(config["SOURCE_PATTERN"])
    if not source_files:
        return None, None
    
    print("📥 Загрузка файлов...")
    all_data = []
    
    for src_file in tqdm(source_files, desc="📂 Обработка", unit="файл", ascii=False, ncols=80):
        df_temp = load_excel_file(src_file, config)
        if df_temp is not None and not df_temp.empty:
            all_data.append(df_temp)
    
    if not all_data:
        print("❌ Не удалось загрузить данные из файлов")
        return None, None
    
    # Объединение данных
    df = pd.concat(all_data, ignore_index=True, sort=False)
    print(f"📊 Всего строк объединено: {len(df):,}")
    
    # Применение фильтров ПЕРЕД валидацией
    df_filtered, filter_stats = apply_filters(df, config)
    
    # Валидация и очистка
    df_clean = validate_and_clean_data(df_filtered)
    
    return df_clean, filter_stats

def calculate_report(df):
    """Выполняет агрегацию данных и расчет отчета."""
    print("\n📊 Расчет показателей...")
    
    # Создаем копию для безопасности
    df_clean = df.copy()
    
    # Начисления и литры (только положительные начисления)
    positive_bonus = df_clean[df_clean["bonus_plus"] > 0].copy()
    negative_bonus = df_clean[df_clean["bonus_minus"] < 0].copy()
    
    # Группировка по периоду - начисления и литры с бонусами
    report = positive_bonus.groupby("period").agg({
        "bonus_plus": "sum",
        "liters": "sum"
    }).rename(columns={
        "bonus_plus": "Бонусов начислено",
        "liters": "Продано литров с начислением бонусов"
    })
    
    # Продано литров всего (включая без бонусов)
    total_liters = df_clean.groupby("period")["liters"].sum()
    report["Продано литров всего"] = total_liters
    
    # Списания (топливо + сопутка)
    if not negative_bonus.empty:
        total_minus = negative_bonus.groupby("period")["bonus_minus"].sum().abs()
        report["Бонусов списано"] = total_minus
    else:
        report["Бонусов списано"] = 0
    
    # Заполнение отсутствующих значений
    report = report.fillna(0)
    
    # Расчет на 1 литр
    report["На 1 литр начислено бонусов"] = report.apply(
        lambda row: row["Бонусов начислено"] / row["Продано литров с начислением бонусов"] 
        if row["Продано литров с начислением бонусов"] != 0 else 0,
        axis=1
    )
    
    # Форматирование периода с русскими названиями месяцев
    try:
        report.index = report.index.strftime("%B %Y")
    except:
        report.index = report.index.strftime("%B %Y")
    
    report = report.reset_index()
    report = report.rename(columns={'period': 'Период'})
    
    return report

def get_sheet_name_from_data(df_report):
    """Генерирует название листа."""
    if len(df_report) == 0:
        return "Отчет"
    
    period_col = "Период"
    if period_col not in df_report.columns:
        if "period" in df_report.columns:
            period_col = "period"
        elif "index" in df_report.columns:
            period_col = "index"
        else:
            return "Отчет"
    
    first_period = df_report[period_col].iloc[0]
    last_period = df_report[period_col].iloc[-1]
    
    sheet_name = f"Отчет за {first_period} - {last_period}"
    
    if len(sheet_name) > 31:
        first_month = first_period.split()[0]
        last_month = last_period.split()[0]
        last_year = last_period.split()[1]
        sheet_name = f"{first_month[:3]}-{last_month[:3]} {last_year}"
    
    return sheet_name

def format_excel_file(file_path, sheet_name, config):
    """Применяет форматирование Excel."""
    try:
        wb = load_workbook(file_path)
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(f"⚠️  Лист '{sheet_name}' не найден, используется активный лист")
            ws = wb.active
            ws.title = sheet_name[:31]
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value or ""))
                    max_length = max(max_length, cell_length)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Форматирование заголовков
        for cell in ws[1]:
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal='center',
                vertical='center',
                shrink_to_fit=True
            )
            cell.font = cell.font.copy(bold=True)
        
        # Финансовый формат для числовых колонок
        financial_cols = {
            'B': 'Бонусов начислено', 
            'C': 'Продано литров с начислением бонусов',
            'D': 'Продано литров всего',
            'E': 'Бонусов списано'
        }
        for col_letter, col_name in financial_cols.items():
            if col_letter in ws.column_dimensions:
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    cell.number_format = config["NUMBER_FORMATS"]["financial"]
        
        # Формат для колонки F
        if 'F' in ws.column_dimensions:
            for row in range(2, ws.max_row + 1):
                cell = ws[f"F{row}"]
                cell.number_format = config["NUMBER_FORMATS"]["rate"]
        
        wb.save(file_path)
        print("✅ Форматирование применено успешно")
        
    except Exception as e:
        print(f"❌ Ошибка форматирования: {e}")

def create_backup(file_path):
    """Создает резервную копию файла."""
    if Path(file_path).exists():
        backup_name = f"{Path(file_path).stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{Path(file_path).suffix}"
        backup_path = Path(file_path).parent / backup_name
        try:
            Path(file_path).rename(backup_path)
            print(f"📦 Создана резервная копия: {backup_name}")
            return True
        except Exception as e:
            print(f"⚠️  Не удалось создать резервную копию: {e}")
    return False

def main():
    """Основная функция выполнения скрипта."""
    start_time = time.time()
    
    print("=" * 70)
    print("🚀 ФОРМИРОВАНИЕ ОТЧЕТА ПО ОБОРОТАМ БОНУСОВ")
    print("=" * 70)
    
    # Установка локали
    setup_locale()
    
    # Проверка исходных файлов
    if not find_source_files(CONFIG["SOURCE_PATTERN"]):
        print("❌ Программа завершена из-за отсутствия исходных файлов")
        input("Нажмите Enter для выхода...")
        sys.exit(1)
    
    # Загрузка и обработка данных (с применением фильтров)
    df_raw, filter_stats = load_and_process_data(CONFIG)
    if df_raw is None:
        print("❌ Не удалось загрузить данные")
        input("Нажмите Enter для выхода...")
        sys.exit(1)
    
    # Расчет отчета
    df_report = calculate_report(df_raw)
    if df_report.empty:
        print("❌ Не удалось рассчитать отчет")
        input("Нажмите Enter для выхода...")
        sys.exit(1)
    
    print(f"📊 Рассчитано периодов: {len(df_report)}")
    
    # Генерация названия листа
    sheet_name = get_sheet_name_from_data(df_report)
    print(f"📋 Название листа: '{sheet_name}'")
    
    # Проверка и создание резервной копии
    dst_path = Path(CONFIG["DST_FILE"])
    if dst_path.exists():
        create_backup(CONFIG["DST_FILE"])
    
    # Сохранение в Excel
    print(f"\n📝 Сохранение в '{CONFIG['DST_FILE']}'...")
    try:
        with pd.ExcelWriter(
            CONFIG["DST_FILE"],
            engine="openpyxl",
            mode='w'
        ) as writer:
            df_report.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )
            
            if len(df_raw) < 10000:
                df_raw_sample = df_raw.head(1000).copy()
                df_raw_sample["date"] = df_raw_sample["date"].dt.strftime("%d.%m.%Y %H:%M")
                df_raw_sample.to_excel(
                    writer,
                    sheet_name="Сырые данные",
                    index=False
                )
        
        print("✅ Файл успешно сохранен")
        
        # Применение форматирования
        format_excel_file(CONFIG["DST_FILE"], sheet_name, CONFIG)
        
        # Итоговая статистика
        print("\n" + "=" * 70)
        print("📊 ИТОГОВАЯ СТАТИСТИКА")
        print("=" * 70)
        print(f"📄 Файл: {CONFIG['DST_FILE']}")
        print(f"📋 Лист: {sheet_name}")
        print(f"📊 Периодов в отчете: {len(df_report)}")
        
        if filter_stats:
            print(f"\n🔍 Фильтрация:")
            print(f"   • Исключено строк: {filter_stats.get('filtered_rows', 0):,}")
            print(f"   • По маркам топлива: {filter_stats.get('filtered_fuel', 0):,}")
            print(f"   • По номерам АЗС: {filter_stats.get('filtered_azs', 0):,}")
        
        print(f"\n📈 Всего начислено бонусов: {df_report['Бонусов начислено'].sum():,.2f}")
        print(f"📉 Всего списано бонусов: {df_report['Бонусов списано'].sum():,.2f}")
        print(f"⛽ Продано литров (с бонусами): {df_report['Продано литров с начислением бонусов'].sum():,.2f}")
        print(f"⛽ Продано литров (всего): {df_report['Продано литров всего'].sum():,.2f}")
        
        total_bonus = df_report['Бонусов начислено'].sum()
        total_liters_with_bonus = df_report['Продано литров с начислением бонусов'].sum()
        if total_liters_with_bonus > 0:
            avg_rate = total_bonus / total_liters_with_bonus
            print(f"🧮 Средний показатель на 1 литр: {avg_rate:,.8f}")
        
        print(f"⏱️  Время выполнения: {time.time() - start_time:.2f} сек")
        print("=" * 70)
        
        # Открываем файл
        open_file_in_default_app(CONFIG["DST_FILE"])
        
    except Exception as e:
        print(f"❌ Ошибка при сохранении файла: {e}")
        print("Подсказка: Закройте файл Excel, если он открыт")
    
    input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Программа прервана пользователем")
    except Exception as e:
        print(f"\n❌ Непредвиденная ошибка: {e}")
        input("Нажмите Enter для выхода...")